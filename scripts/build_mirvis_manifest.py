#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import tempfile
from collections.abc import Iterable
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - helper script
    raise SystemExit("openpyxl is required to run this script") from exc


REPO_URL = "https://github.com/mmirvis/Cell-Anatomy-Scoping-Review.git"
REPO_COMMIT = "a2218f4e65818f3187f2b3c050690bd16fd21c6e"

BYTES_PER_UNIT = {
    "KB": 10**3,
    "MB": 10**6,
    "GB": 10**9,
    "TB": 10**12,
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.replace("\xa0", " ").split())
    return str(value)


def unique_headers(values: Iterable[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        header = clean(value) or f"unnamed_{index}"
        count = seen.get(header, 0) + 1
        seen[header] = count
        headers.append(header if count == 1 else f"{header}_{count}")
    return headers


def sheet_records(path: Path, sheet_name: str, header_row: int = 1, start_row: int | None = None) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    header_values = next(
        worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )
    headers = unique_headers(header_values)
    first_data_row = start_row or header_row + 1

    records: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=first_data_row, values_only=True):
        if not any(value is not None and clean(value) for value in row):
            continue
        record = {header: clean(value) for header, value in zip(headers, row)}
        records.append(record)
    return records


def clone_repo(repo_url: str, commit: str) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="mirvis-manifest-"))
    repo_dir = temp_dir / "repo"
    subprocess.run(
        ["git", "clone", "--depth", "1", repo_url, str(repo_dir)],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    subprocess.run(
        ["git", "-C", str(repo_dir), "checkout", commit],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    return repo_dir


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def extract_urls(*texts: str) -> list[str]:
    def normalize_url(url: str) -> str:
        normalized = url.strip().rstrip(".)],;:")
        normalized = re.sub(r"\.(Source|Sourcedata)$", "", normalized, flags=re.I)
        return normalized

    urls: list[str] = []
    text = " ".join(part for part in texts if part)
    text = re.sub(r"(?<=\w)\s*\.\s*(?=\w)", ".", text)
    text = text.replace("doi.org/", "doi.org/")
    text = text.replace("w ww.", "www.")

    for match in re.findall(r"https?://[^\s)\];,]+", text):
        urls.append(normalize_url(match))

    for match in re.findall(r"10\.6019/EMPIAR-\d+", text, flags=re.I):
        accession = match.split("/")[-1].upper()
        urls.append(f"https://www.ebi.ac.uk/empiar/{accession}/")

    for match in re.findall(r"EMPIAR-\d+", text, flags=re.I):
        accession = match.upper()
        urls.append(f"https://www.ebi.ac.uk/empiar/{accession}/")

    for match in re.findall(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", text, flags=re.I):
        if "EMPIAR-" in match.upper():
            continue
        normalized = normalize_url(match)
        urls.append(f"https://doi.org/{normalized}")

    deduped: list[str] = []
    seen: set[str] = set()
    for url in urls:
        normalized = normalize_url(url.replace(" ", ""))
        if normalized not in seen:
            seen.add(normalized)
            deduped.append(normalized)
    return deduped


def repository_types(urls: Iterable[str]) -> list[str]:
    kinds: list[str] = []
    for url in urls:
        lower = url.lower()
        if "empiar" in lower:
            kinds.append("EMPIAR")
        elif "biostudies" in lower or "bioimages" in lower:
            kinds.append("BioStudies")
        elif "figshare" in lower:
            kinds.append("figshare")
        elif "zenodo" in lower:
            kinds.append("Zenodo")
        elif "webknossos" in lower:
            kinds.append("webKnossos")
        elif "mpi-cbg.de" in lower:
            kinds.append("MPI-CBG Cloud")
        elif "betaseg.github.io" in lower:
            kinds.append("BetaSeg")
        elif "github.com/mobie" in lower:
            kinds.append("GitHub/MoBIE")
        else:
            kinds.append(url.split("/")[2] if "://" in url else "other")
    seen: set[str] = set()
    ordered: list[str] = []
    for kind in kinds:
        if kind not in seen:
            seen.add(kind)
            ordered.append(kind)
    return ordered


def parse_int(text: str) -> int | None:
    if not text:
        return None
    match = re.search(r"\d+", text.replace(",", ""))
    return int(match.group(0)) if match else None


def parse_size_range(raw_size: str, number_of_files: str) -> tuple[str, str, str]:
    if not raw_size:
        return "", "", ""

    text = raw_size.replace(",", "")
    note = ""

    if re.fullmatch(r"\d+(\.\d+)?", text):
        value = float(text) * BYTES_PER_UNIT["GB"]
        note = "numeric cell without units in workbook; interpreted as GB from context"
        return str(int(value)), str(int(value)), note

    if "each" in text.lower():
        file_count = parse_int(number_of_files) or 1
        range_match = re.search(r"([0-9]*\.?[0-9]+)\s*-\s*([0-9]*\.?[0-9]+)\s*(KB|MB|GB|TB)", text, flags=re.I)
        if range_match:
            minimum = float(range_match.group(1)) * BYTES_PER_UNIT[range_match.group(3).upper()] * file_count
            maximum = float(range_match.group(2)) * BYTES_PER_UNIT[range_match.group(3).upper()] * file_count
            note = f"range multiplied by reported file count ({file_count})"
            return str(int(minimum)), str(int(maximum)), note

    matches = re.findall(r"([0-9]*\.?[0-9]+)\s*(KB|MB|GB|TB)", text, flags=re.I)
    if matches:
        total = sum(float(number) * BYTES_PER_UNIT[unit.upper()] for number, unit in matches)
        return str(int(total)), str(int(total)), note

    return "", "", "unable to parse size"


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_study_manifest(repo_dir: Path) -> list[dict[str, str]]:
    includes = sheet_records(
        repo_dir / "Additional file 2_includes_borderlines.xlsx",
        "Data Extraction_Includes",
    )
    borderlines = sheet_records(
        repo_dir / "Additional file 2_includes_borderlines.xlsx",
        "Data Extraction_Borderlines",
    )

    rows: list[dict[str, str]] = []

    for row in includes:
        rows.append(
            {
                "study_id": row["Study ID"],
                "study_slug": slugify(row["Study ID"]),
                "included_status": "included",
                "title": row["Title"],
                "pmid": row["PMID"],
                "year_published": row["Year Published"],
                "journal_published": row["Journal Published"],
                "article_type_general": row["Article Type (General)"],
                "article_type_specific": row["Article Type (Specific)"],
                "open_access": row["Open Access"],
                "data_available": row["Data Available"],
                "data_availability_notes": row["Data Availability Notes"],
                "imaging_modality": row["Imaging Modality"],
                "resolution": row["Resolution"],
                "sample_size_number": row["Sample Size Number"],
                "sample_size_notes": row["Sample Size Notes"],
                "model_system": row["Model System"],
                "model_system_specific": row["Model System Specific"],
                "organelles": row['"Organelles"'],
                "quantifications": row["Quantifications"],
                "comparators_conditions": row["Comparators/Conditions"],
                "comments": row["Comments"],
                "source_file": "Additional file 2_includes_borderlines.xlsx",
                "source_sheet": "Data Extraction_Includes",
                "source_repo_url": REPO_URL,
                "source_commit": REPO_COMMIT,
            }
        )

    for row in borderlines:
        rows.append(
            {
                "study_id": row["Study ID"],
                "study_slug": slugify(row["Study ID"]),
                "included_status": "borderline",
                "title": row["Title"],
                "pmid": row["PMID"],
                "year_published": row["Year Published"],
                "journal_published": row["Journal Published"],
                "article_type_general": "",
                "article_type_specific": "",
                "open_access": "",
                "data_available": "",
                "data_availability_notes": "",
                "imaging_modality": "",
                "resolution": "",
                "sample_size_number": "",
                "sample_size_notes": "",
                "model_system": "",
                "model_system_specific": "",
                "organelles": "",
                "quantifications": "",
                "comparators_conditions": "",
                "comments": row["Reason"] or row["unnamed_6"],
                "source_file": "Additional file 2_includes_borderlines.xlsx",
                "source_sheet": "Data Extraction_Borderlines",
                "source_repo_url": REPO_URL,
                "source_commit": REPO_COMMIT,
            }
        )

    rows.sort(key=lambda row: (row["included_status"], row["study_id"]))
    return rows


def build_public_assets_manifest(repo_dir: Path) -> list[dict[str, str]]:
    rows = sheet_records(
        repo_dir / "Additional file 3_publicdatasets.xlsx",
        "Available datasets",
    )

    assets: list[dict[str, str]] = []
    for row in rows:
        urls = extract_urls(row["Public dataset link"], row["Data Availability Notes"])
        min_bytes, max_bytes, size_note = parse_size_range(row["Total size"], row["Number of files"])
        assets.append(
            {
                "study_id": row["Study ID"],
                "study_slug": slugify(row["Study ID"]),
                "included_status": row["Included/Borderline"].lower(),
                "availability_scope_note": row["unnamed_1"],
                "public_dataset_link_raw": row["Public dataset link"],
                "public_locator_urls": "; ".join(urls),
                "repository_types": "; ".join(repository_types(urls)),
                "imaging_modality": row["Imaging Modality"],
                "resolution": row["Resolution"],
                "sample_size_number": row["Sample Size Number"],
                "sample_size_notes": row["Sample Size Notes"],
                "number_of_files": row["Number of files"],
                "data_format": row["Data format"],
                "file_type": row["File type"],
                "total_size_raw": row["Total size"],
                "total_size_bytes_min": min_bytes,
                "total_size_bytes_max": max_bytes,
                "size_parse_note": size_note,
                "model_system_specific": row["Model System Specific"],
                "organelles": row['"Organelles"'],
                "quantifications": row["Quantifications"],
                "comparators_conditions": row["Comparators/Conditions"],
                "data_availability_notes": row["Data Availability Notes"],
                "source_file": "Additional file 3_publicdatasets.xlsx",
                "source_sheet": "Available datasets",
                "source_repo_url": REPO_URL,
                "source_commit": REPO_COMMIT,
            }
        )

    assets.sort(key=lambda row: (row["included_status"], row["study_id"]))
    return assets


def build_corpus_locator(repo_dir: Path, study_rows: list[dict[str, str]], public_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    datasets = sheet_records(
        repo_dir / "Additional file 5_figure_data.xlsx",
        "Datasets_Mod_Cell_SS",
    )
    study_index = {row["study_id"]: row for row in study_rows}
    public_index = {row["study_id"]: row for row in public_rows}

    locator_rows: list[dict[str, str]] = []
    for index, row in enumerate(datasets, start=1):
        study_id = row["Study ID"]
        study = study_index.get(study_id, {})
        public = public_index.get(study_id, {})

        locator_rows.append(
            {
                "dataset_locator_id": f"{slugify(study_id)}-{index:03d}",
                "study_id": study_id,
                "study_slug": slugify(study_id),
                "included_status": study.get("included_status", "included"),
                "year": row["Year"],
                "journal_published": study.get("journal_published", ""),
                "open_access": study.get("open_access", ""),
                "study_data_available": study.get("data_available", ""),
                "public_locator_urls": public.get("public_locator_urls", ""),
                "public_repository_types": public.get("repository_types", ""),
                "public_total_size_raw": public.get("total_size_raw", ""),
                "public_total_size_bytes_min": public.get("total_size_bytes_min", ""),
                "public_total_size_bytes_max": public.get("total_size_bytes_max", ""),
                "public_size_parse_note": public.get("size_parse_note", ""),
                "data_availability_notes": study.get("data_availability_notes", ""),
                "imaging_modality": row["Imaging Modality"],
                "modality_code": row["Modality code"],
                "xy_nm": row["xy (nm)"],
                "z_nm": row["z (nm)"],
                "organism_type": row["Organism type"],
                "cell_type": row["Cell Type"],
                "cell_type_code": row["Cell Type Code"],
                "abbreviation": row["abbreviation"],
                "min_sample_size": row["Min Sample Size"],
                "sample_size_certain": row["Sample Size Certain?"],
                "sample_size_notes": row["Sample Size Notes"],
                "organelles_common": row["Organelles (common)"],
                "organelles_specialized": row["Organelles (specialized)"],
                "source_file": "Additional file 5_figure_data.xlsx",
                "source_sheet": "Datasets_Mod_Cell_SS",
                "source_repo_url": REPO_URL,
                "source_commit": REPO_COMMIT,
            }
        )

    locator_rows.sort(key=lambda row: (row["included_status"], row["study_id"], row["cell_type"]))
    return locator_rows


def build_summary(study_rows: list[dict[str, str]], public_rows: list[dict[str, str]], locator_rows: list[dict[str, str]]) -> dict[str, Any]:
    included_studies = sum(1 for row in study_rows if row["included_status"] == "included")
    borderline_studies = sum(1 for row in study_rows if row["included_status"] == "borderline")
    included_public = [row for row in public_rows if row["included_status"] == "included"]
    borderline_public = [row for row in public_rows if row["included_status"] == "borderline"]

    def total(rows: list[dict[str, str]], field: str) -> int:
        return sum(int(row[field]) for row in rows if row[field])

    return {
        "source_repo_url": REPO_URL,
        "source_commit": REPO_COMMIT,
        "study_manifest_rows": len(study_rows),
        "included_study_rows": included_studies,
        "borderline_study_rows": borderline_studies,
        "public_asset_rows": len(public_rows),
        "included_public_asset_rows": len(included_public),
        "borderline_public_asset_rows": len(borderline_public),
        "dataset_locator_rows": len(locator_rows),
        "public_size_bytes_min_including_borderline": total(public_rows, "total_size_bytes_min"),
        "public_size_bytes_max_including_borderline": total(public_rows, "total_size_bytes_max"),
        "public_size_bytes_min_included_only": total(included_public, "total_size_bytes_min"),
        "public_size_bytes_max_included_only": total(included_public, "total_size_bytes_max"),
    }


def write_summary_markdown(path: Path, summary: dict[str, Any]) -> None:
    def decimal_tb(value: int) -> str:
        return f"{value / 1e12:.3f} TB"

    body = f"""# Mirvis Corpus Manifest

This directory contains machine-readable manifests derived from the upstream repository:

- repo: {summary["source_repo_url"]}
- commit: `{summary["source_commit"]}`

Files:

- `study_manifest.csv` - included and borderline study-level records
- `public_data_assets.csv` - public dataset locations, links, formats, and reported sizes
- `corpus_locator.csv` - dataset-level locator table joined with study and public-asset metadata
- `manifest_summary.json` - machine-readable summary counts

Regenerate:

- `python scripts/build_mirvis_manifest.py`

Current counts:

- included studies: {summary["included_study_rows"]}
- borderline studies: {summary["borderline_study_rows"]}
- public asset rows: {summary["public_asset_rows"]}
- dataset locator rows: {summary["dataset_locator_rows"]}

Reported public-data footprint:

- included studies only: {decimal_tb(summary["public_size_bytes_min_included_only"])} to {decimal_tb(summary["public_size_bytes_max_included_only"])}
- including the one borderline public row: {decimal_tb(summary["public_size_bytes_min_including_borderline"])} to {decimal_tb(summary["public_size_bytes_max_including_borderline"])}

Notes:

- These totals come from the sizes reported in `Additional file 3_publicdatasets.xlsx`.
- Some size fields are ambiguous in the source workbook. The CSV preserves both the raw size string and any parse note.
- `corpus_locator.csv` is the most useful operational table when the question is "what dataset exists and where does its public data live?"
"""
    path.write_text(body, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build local manifests from the Mirvis corpus repo.")
    parser.add_argument(
        "--output-dir",
        default="references/manifests",
        help="Directory to write generated manifest files into.",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    repo_dir = clone_repo(REPO_URL, REPO_COMMIT)

    study_rows = build_study_manifest(repo_dir)
    public_rows = build_public_assets_manifest(repo_dir)
    locator_rows = build_corpus_locator(repo_dir, study_rows, public_rows)
    summary = build_summary(study_rows, public_rows, locator_rows)

    write_csv(output_dir / "study_manifest.csv", study_rows)
    write_csv(output_dir / "public_data_assets.csv", public_rows)
    write_csv(output_dir / "corpus_locator.csv", locator_rows)
    (output_dir / "manifest_summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    write_summary_markdown(output_dir / "README.md", summary)


if __name__ == "__main__":
    main()
